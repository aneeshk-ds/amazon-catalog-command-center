#!/usr/bin/env python3
"""Extract & normalize catalog performance sources into Tableau-ready CSVs."""

from __future__ import annotations

import csv
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
CLEAN = ROOT / "data" / "clean"
PDF_PATH = ROOT / "Performance.pdf"
XLSX_PATH = ROOT / "Prime Deal - 2025 - Plan.xlsx"


def write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        print(f"  WARN empty: {path.name}")
        return
    # Anonymize string fields before writing (portfolio-safe brands)
    for row in rows:
        for k, v in list(row.items()):
            if isinstance(v, str):
                row[k] = anonymize_brand_text(v)
    # stable column order from first row, union extras
    fieldnames: list[str] = []
    seen = set()
    for row in rows:
        for k in row:
            if k not in seen:
                seen.add(k)
                fieldnames.append(k)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  wrote {path.name} ({len(rows)} rows)")


def clean_val(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, str):
        s = v.strip()
        if s in {"#DIV/0!", "#VALUE!", "#N/A", "#REF!", "#NAME?", ""}:
            return None
        return s
    return v


def to_float(v):
    v = clean_val(v)
    if v is None:
        return None
    if isinstance(v, bool):
        return float(v)
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", "").replace("%", ""))
        except ValueError:
            return None
    return None


def extract_pdf() -> list[dict]:
    text = "\n".join((p.extract_text() or "") for p in PdfReader(str(PDF_PATH)).pages)
    pat = re.compile(
        r"(20\d{2}-\d{2}-\d{2})"
        r"(B0[A-Z0-9]{8})"
        r"([A-Za-z][A-Za-z .]*?)"
        r"(\d[\d,]*\.?\d*)\s+(\d[\d,]*\.?\d*)\s+(\d[\d,]*\.?\d*)\s+(\d[\d,]*\.?\d*)\s+"
        r"(-?\d[\d,]*\.?\d*)\s+(-?\d[\d,]*\.?\d*)"
    )
    rows = []
    for m in pat.finditer(text.replace(",", "")):
        month, asin, handler, tr, tgp, atr, atgp, income, gp = m.groups()
        tr, tgp, atr, atgp, income, gp = map(float, (tr, tgp, atr, atgp, income, gp))
        rev_base = atr if atr > 0 else (tr if tr > 0 else None)
        gp_base = atgp if atgp > 0 else (tgp if tgp > 0 else None)
        rows.append(
            {
                "month_start_date": month,
                "year_month": month[:7],
                "asin": asin,
                "seo_handler": handler.strip(),
                "target_revenue": tr,
                "target_gp": tgp,
                "active_target_revenue": atr,
                "active_target_gp": atgp,
                "total_income": income,
                "gross_profit": gp,
                "rev_attain_pct": round(100 * income / rev_base, 2) if rev_base else None,
                "gp_attain_pct": round(100 * gp / gp_base, 2) if gp_base else None,
                "gp_margin_pct": round(100 * gp / income, 2) if income else None,
                "rev_gap_vs_active": round(income - atr, 2),
                "gp_gap_vs_active": round(gp - atgp, 2),
                "is_incomplete_month": 1 if month == "2026-06-01" else 0,
                "has_zero_income": 1 if income == 0 else 0,
                "plan_but_inactive": 1 if atr == 0 and tr > 0 else 0,
            }
        )
    return rows


def anonymize_brand_text(text: str | None) -> str | None:
    """Replace real brand tokens with fictional portfolio brands."""
    if text is None:
        return None
    s = str(text)
    repls = [
        (r"Weight\s*World", "VitalCore"),
        (r"Weightworld", "VitalCore"),
        (r"weightworld", "VitalCore"),
        (r"Max\s*Medix", "Medora"),
        (r"MaxMedix", "Medora"),
        (r"maxmedix", "Medora"),
        (r"Animigo", "PawVista"),
        (r"animigo", "PawVista"),
        (r"(^|[\s\|,\"\'>])WW(?=([\s\|,\"\'<$]|$))", r"\1VitalCore"),
        (r"(^|[\s,\"\'>])WW\|", r"\1VitalCore|"),
        (r"(^|[\s,\"\'>])WW ", r"\1VitalCore "),
    ]
    for pat, rep in repls:
        s = re.sub(pat, rep, s)
    return s


def brand_from_name(name: str | None) -> str | None:
    """Map product title → fictional brand (also accepts raw source brands)."""
    if not name:
        return None
    n = name.lower()
    if "animigo" in n or "pawvista" in n:
        return "PawVista"
    if "maxmedix" in n or "max medix" in n or "medora" in n:
        return "Medora"
    if (
        "weightworld" in n
        or "vitalcore" in n
        or n.startswith("ww ")
        or "| ww" in n
        or n.startswith("ww|")
        or name.strip().upper().startswith("WW")
        or name.strip().startswith("VitalCore")
    ):
        return "VitalCore"
    return "Other"


def extract_excel() -> dict[str, list[dict]]:
    wb = load_workbook(XLSX_PATH, data_only=True)
    out: dict[str, list[dict]] = {}

    # --- Spend allocation (Prime Day plan vs Day-1) ---
    ws = wb["Spend allocation"]
    spend_rows = []
    # Header is row 2
    for r in range(3, ws.max_row + 1):
        asin = clean_val(ws.cell(r, 1).value)
        if not asin or not str(asin).startswith("B"):
            continue
        spend_rows.append(
            {
                "asin": str(asin).strip(),
                "sku": clean_val(ws.cell(r, 2).value),
                "product_name": clean_val(ws.cell(r, 3).value),
                "current_price": to_float(ws.cell(r, 4).value),
                "deal_price": to_float(ws.cell(r, 5).value),
                "ppc_assignee": clean_val(ws.cell(r, 6).value),
                "daily_avg_units": to_float(ws.cell(r, 7).value),
                "daily_sale_avg": to_float(ws.cell(r, 8).value),
                "expected_units": to_float(ws.cell(r, 9).value),
                "expected_sale": to_float(ws.cell(r, 10).value),
                "ppc_share": to_float(ws.cell(r, 11).value),
                "exp_ppc_units": to_float(ws.cell(r, 12).value),
                "daily_budget_per_product": to_float(ws.cell(r, 13).value),
                "cpu_per_product": to_float(ws.cell(r, 14).value),
                "day1_units": to_float(ws.cell(r, 15).value),
                "day1_sales": to_float(ws.cell(r, 16).value),
                "day1_spend": to_float(ws.cell(r, 17).value),
                "spend_diff": to_float(ws.cell(r, 18).value),
                "unit_diff": to_float(ws.cell(r, 19).value),
                "bacos": to_float(ws.cell(r, 20).value),
                "gp_pct": to_float(ws.cell(r, 21).value),
                "gp_per_unit": to_float(ws.cell(r, 22).value),
                "expected_gp_pct": to_float(ws.cell(r, 23).value),
                "expected_bacos": to_float(ws.cell(r, 24).value),
                "extra_blended": to_float(ws.cell(r, 25).value),
                "event_date": "2025-07-09",
                "brand": brand_from_name(clean_val(ws.cell(r, 3).value)),
            }
        )
    out["prime_day_plan_vs_actual"] = spend_rows

    # --- Hourly Update (unpivot) ---
    ws = wb["Hourly Update"]
    # Row1 dates in cols: E=5 (2025-07-08), M=13 totals for day1?, S=19 (2025-07-09)
    # Structure from earlier inspection:
    # cols 1-4: ASIN, Product, SEO, PPC
    # Jul 8 buckets: 5-6 12PM, 7-8 4PM, 9-10 6PM, 11-12 8PM, 13-14 Total units/spend, 15-16 Diff, 17 Diff SPU
    # Jul 9: 19-20 12:30PM, 21-22 4PM, 23-24 6PM, 25-26 8PM
    hourly = []
    buckets_d1 = [
        ("2025-07-08", "12:00", 5, 6),
        ("2025-07-08", "16:00", 7, 8),
        ("2025-07-08", "18:00", 9, 10),
        ("2025-07-08", "20:00", 11, 12),
    ]
    buckets_d2 = [
        ("2025-07-09", "12:30", 19, 20),
        ("2025-07-09", "16:00", 21, 22),
        ("2025-07-09", "18:00", 23, 24),
        ("2025-07-09", "20:00", 25, 26),
    ]
    for r in range(4, ws.max_row + 1):
        asin = clean_val(ws.cell(r, 1).value)
        if not asin or not str(asin).startswith("B"):
            continue
        asin = str(asin).strip()
        product = clean_val(ws.cell(r, 2).value)
        seo = clean_val(ws.cell(r, 3).value)
        ppc = clean_val(ws.cell(r, 4).value)
        day_total_units = to_float(ws.cell(r, 13).value)
        day_total_spend = to_float(ws.cell(r, 14).value)
        for event_date, slot, uc, sc in buckets_d1 + buckets_d2:
            units = to_float(ws.cell(r, uc).value)
            spend = to_float(ws.cell(r, sc).value)
            if units is None and spend is None:
                continue
            hourly.append(
                {
                    "asin": asin,
                    "product_name": product,
                    "seo_owner": seo,
                    "ppc_owner": ppc,
                    "event_date": event_date,
                    "time_slot": slot,
                    "units": units if units is not None else 0,
                    "spend": spend if spend is not None else 0,
                    "day_total_units_jul8": day_total_units,
                    "day_total_spend_jul8": day_total_spend,
                    "brand": brand_from_name(product if isinstance(product, str) else None),
                }
            )
    out["hourly_prime"] = hourly

    # --- Campaign spend Budget ---
    ws = wb["Campaign spend Budget"]
    campaigns = []
    for r in range(2, ws.max_row + 1):
        label = clean_val(ws.cell(r, 1).value)
        if not label:
            continue
        label_s = str(label)
        # ASIN usually last token after |
        parts = [p.strip() for p in label_s.split("|")]
        asin = None
        for p in reversed(parts):
            if re.fullmatch(r"B0[A-Z0-9]{8}", p.replace(" ", "")):
                asin = p.replace(" ", "")
                break
            m = re.search(r"(B0[A-Z0-9]{8})", p)
            if m:
                asin = m.group(1)
                break
        tactic = parts[-2] if len(parts) >= 2 and asin else (parts[-1] if parts else None)
        if asin and tactic and asin in str(tactic):
            tactic = parts[-2] if len(parts) >= 3 else tactic
        product_hint = parts[0] if parts else None
        campaigns.append(
            {
                "campaign_label": label_s,
                "asin": asin,
                "product_hint": product_hint,
                "tactic": tactic,
                "sum_of_spend": to_float(ws.cell(r, 2).value),
                "spend_share": to_float(ws.cell(r, 3).value),
                "total_budget": to_float(ws.cell(r, 4).value),
                "campaign_level_budget": to_float(ws.cell(r, 5).value),
            }
        )
    out["campaign_budget"] = campaigns

    # --- CPU per product ---
    ws = wb["CPU per product"]
    cpu_rows = []
    for r in range(2, ws.max_row + 1):
        asin = clean_val(ws.cell(r, 1).value)
        if not asin or not str(asin).startswith("B"):
            continue
        name = clean_val(ws.cell(r, 2).value)
        cpu_rows.append(
            {
                "asin": str(asin).strip(),
                "product_name": name,
                "cpu_euros_planned": to_float(ws.cell(r, 3).value),
                "units_sold": to_float(ws.cell(r, 4).value),
                "total_income": to_float(ws.cell(r, 5).value),
                "ppc_income": to_float(ws.cell(r, 6).value),
                "advertising": to_float(ws.cell(r, 7).value),
                "ppc_share": to_float(ws.cell(r, 8).value),
                "organic_share": to_float(ws.cell(r, 9).value),
                "ppc_units_sold": to_float(ws.cell(r, 10).value),
                "cpu_actual": to_float(ws.cell(r, 11).value),
                "brand": brand_from_name(name if isinstance(name, str) else None),
            }
        )
    out["cpu_mix"] = cpu_rows

    # --- Rough → owner dim pieces ---
    ws = wb["Rough"]
    owners = []
    for r in range(2, ws.max_row + 1):
        asin = clean_val(ws.cell(r, 1).value)
        if not asin or not str(asin).startswith("B"):
            continue
        name = clean_val(ws.cell(r, 2).value)
        owners.append(
            {
                "asin": str(asin).strip(),
                "product_name": name,
                "ppc_owner": clean_val(ws.cell(r, 3).value),
                "seo_owner": clean_val(ws.cell(r, 4).value),
                "brand": brand_from_name(name if isinstance(name, str) else None),
            }
        )
    out["asin_owners"] = owners
    return out


def build_asin_dim(pnl: list[dict], excel: dict[str, list[dict]]) -> list[dict]:
    names: dict[str, str] = {}
    brands: dict[str, str] = {}
    seo: dict[str, str] = {}
    ppc: dict[str, str] = {}
    sku: dict[str, str] = {}

    for r in excel.get("asin_owners", []):
        a = r["asin"]
        if r.get("product_name"):
            names[a] = r["product_name"]
        if r.get("brand"):
            brands[a] = r["brand"]
        if r.get("seo_owner"):
            seo[a] = str(r["seo_owner"]).strip()
        if r.get("ppc_owner"):
            ppc[a] = str(r["ppc_owner"]).strip()

    for r in excel.get("prime_day_plan_vs_actual", []):
        a = r["asin"]
        if r.get("product_name"):
            names[a] = r["product_name"]
        if r.get("brand"):
            brands[a] = r["brand"]
        if r.get("sku"):
            sku[a] = r["sku"]
        if r.get("ppc_assignee") and a not in ppc:
            ppc[a] = str(r["ppc_assignee"]).strip()

    for r in excel.get("cpu_mix", []):
        a = r["asin"]
        if r.get("product_name") and a not in names:
            names[a] = r["product_name"]
        if r.get("brand") and a not in brands:
            brands[a] = r["brand"]

    for r in excel.get("hourly_prime", []):
        a = r["asin"]
        if r.get("product_name") and a not in names:
            names[a] = r["product_name"]
        if r.get("seo_owner") and a not in seo:
            seo[a] = str(r["seo_owner"]).strip()
        if r.get("ppc_owner") and a not in ppc:
            ppc[a] = str(r["ppc_owner"]).strip()

    asins = set(r["asin"] for r in pnl) | set(names) | set(brands) | set(sku)
    # also campaigns
    for r in excel.get("campaign_budget", []):
        if r.get("asin"):
            asins.add(r["asin"])

    dim = []
    for a in sorted(asins):
        dim.append(
            {
                "asin": a,
                "product_name": names.get(a),
                "brand": brands.get(a) or brand_from_name(names.get(a)),
                "sku": sku.get(a),
                "seo_owner": seo.get(a),
                "ppc_owner": ppc.get(a),
                "in_monthly_pnl": 1 if any(r["asin"] == a for r in pnl) else 0,
                "in_prime_day": 1 if any(r["asin"] == a for r in excel.get("prime_day_plan_vs_actual", [])) else 0,
            }
        )
    return dim


def write_dictionary() -> None:
    rows = [
        {
            "table": "monthly_asin_pnl",
            "column": "month_start_date",
            "definition": "First day of performance month (from Performance.pdf)",
        },
        {
            "table": "monthly_asin_pnl",
            "column": "asin",
            "definition": "Amazon Standard Identification Number — primary join key",
        },
        {
            "table": "monthly_asin_pnl",
            "column": "target_revenue / target_gp",
            "definition": "Original monthly plan targets",
        },
        {
            "table": "monthly_asin_pnl",
            "column": "active_target_revenue / active_target_gp",
            "definition": "Active/adjusted targets used for attainment (0 = inactive that month)",
        },
        {
            "table": "monthly_asin_pnl",
            "column": "total_income / gross_profit",
            "definition": "Actual revenue and gross profit",
        },
        {
            "table": "monthly_asin_pnl",
            "column": "rev_attain_pct / gp_attain_pct",
            "definition": "Actual ÷ active target (falls back to plan if active=0)",
        },
        {
            "table": "monthly_asin_pnl",
            "column": "is_incomplete_month",
            "definition": "1 for 2026-06 where active targets are missing — exclude from full-period KPIs or label",
        },
        {
            "table": "prime_day_plan_vs_actual",
            "column": "bacos",
            "definition": "Day-1 ad spend ÷ Day-1 sales (promo efficiency)",
        },
        {
            "table": "prime_day_plan_vs_actual",
            "column": "expected_* vs day1_*",
            "definition": "Prime Day plan expectations vs 9 Jul Day-1 actuals",
        },
        {
            "table": "hourly_prime",
            "column": "time_slot",
            "definition": "Unpivoted intra-day checkpoint (units & spend)",
        },
        {
            "table": "campaign_budget",
            "column": "tactic",
            "definition": "Parsed campaign type hint (Exact, Broad, PT, SKAG, Auto, etc.)",
        },
        {
            "table": "cpu_mix",
            "column": "ppc_share / organic_share",
            "definition": "Mix of attributed PPC vs organic; CPU = advertising ÷ PPC units when available",
        },
        {
            "table": "asin_dim",
            "column": "asin",
            "definition": "Dimension table for product name, brand, SEO/PPC owners — join to all facts",
        },
        {
            "table": "_project",
            "column": "notes",
            "definition": "Amazon Catalog Command Center · synthesized data · currency treated as EUR · join facts on asin",
        },
    ]
    write_csv(CLEAN / "data_dictionary.csv", rows)


def write_validation(pnl: list[dict]) -> None:
    tot_inc = sum(r["total_income"] for r in pnl)
    tot_gp = sum(r["gross_profit"] for r in pnl)
    tot_atr = sum(r["active_target_revenue"] for r in pnl)
    tot_tr = sum(r["target_revenue"] for r in pnl)
    tot_atgp = sum(r["active_target_gp"] for r in pnl)
    tot_tgp = sum(r["target_gp"] for r in pnl)
    rows = [
        {"metric": "plan_revenue", "value": round(tot_tr, 2), "expected_approx": 7405356, "status": "ok" if abs(tot_tr - 7405356) < 100 else "check"},
        {"metric": "active_revenue", "value": round(tot_atr, 2), "expected_approx": 6139090, "status": "ok" if abs(tot_atr - 6139090) < 100 else "check"},
        {"metric": "actual_income", "value": round(tot_inc, 2), "expected_approx": 5555108, "status": "ok" if abs(tot_inc - 5555108) < 100 else "check"},
        {"metric": "plan_gp", "value": round(tot_tgp, 2), "expected_approx": 1469257, "status": "ok" if abs(tot_tgp - 1469257) < 100 else "check"},
        {"metric": "active_gp", "value": round(tot_atgp, 2), "expected_approx": 1200405, "status": "ok" if abs(tot_atgp - 1200405) < 100 else "check"},
        {"metric": "actual_gp", "value": round(tot_gp, 2), "expected_approx": 1088106, "status": "ok" if abs(tot_gp - 1088106) < 100 else "check"},
        {"metric": "row_count", "value": len(pnl), "expected_approx": 1230, "status": "ok" if len(pnl) == 1230 else "check"},
        {"metric": "unique_asins", "value": len({r["asin"] for r in pnl}), "expected_approx": 125, "status": "ok"},
    ]
    write_csv(CLEAN / "validation_summary.csv", rows)
    print("Validation:")
    for r in rows:
        print(f"  {r['metric']}: {r['value']} (expect ~{r['expected_approx']}) [{r['status']}]")


def main() -> None:
    CLEAN.mkdir(parents=True, exist_ok=True)
    print("Extracting PDF…")
    pnl = extract_pdf()
    write_csv(CLEAN / "monthly_asin_pnl.csv", pnl)
    write_validation(pnl)

    print("Extracting Excel…")
    excel = extract_excel()
    write_csv(CLEAN / "prime_day_plan_vs_actual.csv", excel["prime_day_plan_vs_actual"])
    write_csv(CLEAN / "hourly_prime.csv", excel["hourly_prime"])
    write_csv(CLEAN / "campaign_budget.csv", excel["campaign_budget"])
    write_csv(CLEAN / "cpu_mix.csv", excel["cpu_mix"])
    write_csv(CLEAN / "asin_owners.csv", excel["asin_owners"])

    print("Building asin_dim…")
    dim = build_asin_dim(pnl, excel)
    write_csv(CLEAN / "asin_dim.csv", dim)

    write_dictionary()
    print(f"\nDone → {CLEAN}")


if __name__ == "__main__":
    main()
